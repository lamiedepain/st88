from flask import Flask, render_template, jsonify, request, send_file
import openpyxl
from datetime import datetime
import os
import shutil
import calendar
from io import BytesIO

# Groupes pour assignation (mêmes que côté frontend)
GROUPS = {
    'Encadrant': ['DEBREYNE', 'LUTARD', 'HAUTDECOEUR', 'GRENET', 'FOURCADE', 'GONCALVES', 'SIGALA', 'BETTINGER', 'TUCOULET', 'VRBOVSKA'],
    'Surveillant de travaux': ['BOURGOIN', 'MERCADIEU', 'GARCIA', 'LARROUDE', 'SAMITIER', 'PIEL'],
    'Encadrant Propreté': ['GOURVIAT', 'LARTIGUE', 'TRIQUENEAUX', 'ESPERON', 'ROUGLAN', 'NOURRI'],
    'Agents Voirie': ['BERRIO-GAUDNER', 'FONTENEAU', 'GUIJARRO', 'GOUREAU', 'LABORIE', 'LARRIEU', 'LEVIGNAT', 'MARTIN-HERNANDEZ', 'PIERRE', 'SOLA', 'WEISS'],
    'Agent EV': ['DELANDE', 'DA SILVA REIS', 'ELMAGROUD', 'ESTEVE', 'KADRI', 'MALLET', 'MAURY', 'MOINGT', 'REY', 'TADJROUNA', 'VILLENEUVE'],
    'COMMUN - Magasinier': ['VOL', 'GENNA', 'BERNARD', 'HAUBRAICHE']
}

app = Flask(__name__)

EXCEL_FILE = '2026 - PRESENCES_CONGES VOIRIE ESPACES VERTS ST8 (1).xlsx'
BACKUP_DIR = 'backups'


# Normalise les codes de statut venant du fichier Excel
def normalize_status(value):
    if value is None:
        return ''
    try:
        s = str(value).strip()
    except Exception:
        return ''
    # Corriger le cas où le fichier contient 'MA' (confusion) -> traiter comme 'M' (Maladie)
    if s.upper() == 'MA':
        return 'M'
    return s

# Créer une sauvegarde au démarrage
if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)

backup_file = os.path.join(BACKUP_DIR, f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
if os.path.exists(EXCEL_FILE):
    shutil.copy2(EXCEL_FILE, backup_file)
    print(f"✅ Sauvegarde créée: {backup_file}")

@app.route('/')
def index():
    return render_template('agents.html')

@app.route('/planning')
def planning():
    return render_template('planning.html')

@app.route('/generator')
def generator():
    return render_template('generator.html')

@app.route('/api/agents', methods=['GET'])
def get_agents():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb['config']
        
        # Lire les en-têtes (ligne 2)
        headers = [cell.value for cell in sheet[2]]
        
        # Lire les agents (lignes 3 à 75)
        agents = []
        for row_idx in range(3, 76):  # Lignes 3 à 75
            row = [cell.value for cell in sheet[row_idx]]
            if row[5]:  # Si le nom existe (colonne 6 = index 5)
                nom = row[5].strip() if isinstance(row[5], str) else row[5]
                prenom = row[6].strip() if isinstance(row[6], str) else row[6]
                
                agent = {
                    'index': row_idx,  # Numéro de ligne réel
                    'matricule': row[4] or '',
                    'nom': nom or '',
                    'prenom': prenom or '',
                    'data': row
                }
                agents.append(agent)
        
        return jsonify({
            'success': True,
            'headers': headers,
            'agents': agents
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/agents/<int:index>', methods=['PUT'])
def update_agent(index):
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb['config']
        
        # index est déjà le numéro de ligne réel
        row_num = index
        
        # Mettre à jour les cellules - s'assurer que toutes les valeurs sont sérialisables
        for col_idx, value in enumerate(data.get('data', []), start=1):
            # Convertir les valeurs None en chaîne vide
            if value is None:
                value = ''
            sheet.cell(row=row_num, column=col_idx, value=value)
        
        wb.save(EXCEL_FILE)
        return jsonify({'success': True})
    except Exception as e:
        print(f"Erreur lors de la mise à jour de l'agent {index}: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/agents/<int:index>', methods=['DELETE'])
def delete_agent(index):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb['config']
        
        # index est déjà le numéro de ligne réel
        row_num = index
        sheet.delete_rows(row_num, 1)
        
        wb.save(EXCEL_FILE)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/planning/<month>', methods=['GET'])
def get_planning(month):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        # Le mois est au format "Janvier 2026"
        if month not in wb.sheetnames:
            return jsonify({'success': False, 'error': 'Mois introuvable'}), 404
        
        sheet = wb[month]
        
        # Lire toutes les données du planning
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        
        return jsonify({
            'success': True,
            'month': month,
            'data': data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/planning/<month>', methods=['PUT'])
def update_planning(month):
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
            
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        if month not in wb.sheetnames:
            return jsonify({'success': False, 'error': 'Mois introuvable'}), 404
        
        sheet = wb[month]
        
        # Mettre à jour les cellules modifiées
        for update in data.get('updates', []):
            row = update['row']
            col = update['col']
            value = update['value']
            sheet.cell(row=row, column=col, value=value)
        
        wb.save(EXCEL_FILE)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/months', methods=['GET'])
def get_months():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        months = [name for name in wb.sheetnames if name != 'config']
        return jsonify({
            'success': True,
            'months': months
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/planning-data/<year>/<month>', methods=['GET'])
def get_planning_data(year, month):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        # Trouver la feuille correspondante
        month_names = ['Janvier', 'Fevrier', 'Mars', 'Avril', 'Mai', 'Juin',
                      'Juillet', 'Aout', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
        month_idx = int(month) - 1
        sheet_name = f"{month_names[month_idx]} {year}"
        
        if sheet_name not in wb.sheetnames:
            return jsonify({'success': False, 'error': 'Feuille introuvable'}), 404
        
        sheet = wb[sheet_name]
        
        # Structure: ligne 11+ = agents, colonnes 16+ (colonne P) = jours 1 à 31
        # Ligne 10 = en-têtes (Matricule, Nom, Prénom...)
        agents_data = []
        empty_rows_count = 0  # Compteur de lignes vides consécutives
        
        for row_idx in range(11, 100):  # Lignes agents (augmenté pour récupérer tous les agents)
            row = sheet[row_idx]
            matricule = row[0].value
            nom = row[1].value
            prenom = row[2].value
            
            if not nom:  # Si pas de nom
                empty_rows_count += 1
                if empty_rows_count >= 5:  # Arrêter après 5 lignes vides consécutives
                    break
                continue  # Sauter cette ligne et continuer
            
            empty_rows_count = 0  # Réinitialiser le compteur si on trouve un agent
            
            # Nettoyer le nom et prénom (enlever espaces en trop)
            nom = nom.strip() if isinstance(nom, str) else nom
            prenom = prenom.strip() if isinstance(prenom, str) else prenom
            
            # Récupérer les statuts pour chaque jour (colonnes 15 à 45 = index 15 à 45)
            days_status = []
            for col_idx in range(15, 46):  # Colonnes P à AT (31 jours max)
                cell_value = row[col_idx].value
                days_status.append(normalize_status(cell_value))
            
            agents_data.append({
                'matricule': matricule or '',
                'nom': nom or '',
                'prenom': prenom or '',
                'days': days_status
            })
        
        return jsonify({
            'success': True,
            'agents': agents_data
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/generate-week', methods=['POST'])
def generate_week():
    try:
        data = request.json or {}
        week = data.get('week')  # format 'YYYY-Www' from input[type=week]
        group = data.get('group', 'all')
        slots = int(data.get('slots', 1))

        if not week:
            return jsonify({'success': False, 'error': 'week is required'}), 400

        # Parse week string
        if '-W' in week:
            parts = week.split('-W')
            year = int(parts[0])
            week_no = int(parts[1])
        else:
            # fallback: assume format 'YYYY-Www' anyway
            parts = week.split('-')
            year = int(parts[0])
            week_no = int(parts[1].lstrip('W'))

        # Compute dates for the ISO week (Monday..Sunday)
        dates = []
        for weekday in range(1, 8):
            d = datetime.fromisocalendar(year, week_no, weekday)
            dates.append(d)

        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

        # Helper to determine group membership
        def in_group(nom, group_name):
            if not nom:
                return False
            nom_u = str(nom).strip().upper()
            if group_name == 'all':
                return True
            members = GROUPS.get(group_name, [])
            return any(nom_u == m for m in members)

        # statuses considered as absence
        absent_statuses = set(['CA','RTT','CEX','R','M','AT','F','AST','PC','TP','TAD'])

        # Build availability per date
        availability = {}
        pool = []  # unique agents available at least once
        pool_map = {}

        for d in dates:
            month_names = ['Janvier', 'Fevrier', 'Mars', 'Avril', 'Mai', 'Juin',
                          'Juillet', 'Aout', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
            sheet_name = f"{month_names[d.month-1]} {d.year}"
            day = d.day
            available_agents = []

            if sheet_name not in wb.sheetnames:
                availability[d.strftime('%Y-%m-%d')] = []
                continue

            sheet = wb[sheet_name]
            for row_idx in range(11, 100):
                row = sheet[row_idx]
                matricule = row[0].value
                nom = row[1].value
                prenom = row[2].value

                if not nom:
                    continue

                if not in_group(nom, group):
                    continue

                col_idx = 14 + day  # day 1 => index 15 per existing logic
                cell_value = None
                try:
                    cell_value = row[col_idx].value
                except Exception:
                    cell_value = None

                status = normalize_status(cell_value)
                if status.upper() in absent_statuses:
                    is_available = False
                else:
                    # treat empty or 'P' as available
                    if status == '' or status.upper() == 'P' or status.lower() == 'présent' or status.lower() == 'present':
                        is_available = True
                    else:
                        # unknown code -> assume available
                        is_available = True

                if is_available:
                    agent_info = {
                        'matricule': matricule or '',
                        'nom': nom or '',
                        'prenom': prenom or '',
                        'fullName': f"{nom} {prenom}".strip(),
                        'row': row_idx
                    }
                    available_agents.append(agent_info)
                    key = (agent_info['nom'], agent_info['prenom'])
                    if key not in pool_map:
                        pool_map[key] = agent_info
                        pool.append(agent_info)

            availability[d.strftime('%Y-%m-%d')] = available_agents

        # Round-robin assign using the pool but only choosing agents available that day
        assignments = {}
        if len(pool) == 0:
            # no available agents
            for d in dates:
                assignments[d.strftime('%Y-%m-%d')] = {'assigned': [], 'available': availability.get(d.strftime('%Y-%m-%d'), [])}
        else:
            cursor = 0
            for d in dates:
                date_key = d.strftime('%Y-%m-%d')
                avail = availability.get(date_key, [])
                assigned = []
                if avail:
                    # attempt to pick 'slots' distinct agents from pool who are in avail
                    picked = []
                    attempts = 0
                    while len(picked) < slots and attempts < len(pool) * 2:
                        candidate = pool[cursor % len(pool)]
                        cursor += 1
                        attempts += 1
                        # is candidate available today?
                        if any((candidate['nom'] == a['nom'] and candidate['prenom'] == a['prenom']) for a in avail):
                            # avoid duplicates
                            if not any((candidate['nom'] == p['nom'] and candidate['prenom'] == p['prenom']) for p in picked):
                                picked.append(candidate)
                    assigned = picked

                assignments[date_key] = {'assigned': assigned, 'available': availability.get(date_key, [])}

        return jsonify({'success': True, 'assignments': assignments})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/apply-week', methods=['POST'])
def apply_week():
    try:
        data = request.json or {}
        week = data.get('week')
        group = data.get('group', 'all')
        slots = int(data.get('slots', 1))

        if not week:
            return jsonify({'success': False, 'error': 'week is required'}), 400

        # Reuse generate_week logic to build assignments
        # We'll call the internal function by reusing the POST payload flow: generate_week already computes assignments,
        # but to avoid duplicating code we will call generate_week() functionally is complicated; instead recompute here (duplicate of logic)

        if '-W' in week:
            parts = week.split('-W')
            year = int(parts[0])
            week_no = int(parts[1])
        else:
            parts = week.split('-')
            year = int(parts[0])
            week_no = int(parts[1].lstrip('W'))

        dates = [datetime.fromisocalendar(year, week_no, wd) for wd in range(1,8)]

        wb = openpyxl.load_workbook(EXCEL_FILE)

        # create a timestamped backup
        backup_file = os.path.join(BACKUP_DIR, f"apply_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(backup_file)

        # Build availability like generate_week
        absent_statuses = set(['CA','RTT','CEX','R','M','AT','F','AST','PC','TP','TAD'])
        availability = {}
        pool = []
        pool_map = {}

        for d in dates:
            month_names = ['Janvier', 'Fevrier', 'Mars', 'Avril', 'Mai', 'Juin',
                          'Juillet', 'Aout', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
            sheet_name = f"{month_names[d.month-1]} {d.year}"
            day = d.day
            available_agents = []

            if sheet_name not in wb.sheetnames:
                availability[d.strftime('%Y-%m-%d')] = []
                continue

            sheet = wb[sheet_name]
            for row_idx in range(11, 100):
                nom = sheet.cell(row=row_idx, column=2).value
                prenom = sheet.cell(row=row_idx, column=3).value

                if not nom:
                    continue

                nom_u = str(nom).strip().upper()
                # group filter
                if group != 'all':
                    members = GROUPS.get(group, [])
                    if nom_u not in members:
                        continue

                col_idx = 14 + day  # align with read logic: day1 -> column 15 (P)
                cell_value = None
                try:
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                except Exception:
                    cell_value = None

                status = normalize_status(cell_value)
                if status.upper() in absent_statuses:
                    is_available = False
                else:
                    is_available = True

                if is_available:
                    agent_info = {'matricule': sheet.cell(row=row_idx, column=1).value or '', 'nom': nom, 'prenom': prenom, 'fullName': f"{nom} {prenom}".strip(), 'row': row_idx}
                    available_agents.append(agent_info)
                    key = (agent_info['nom'], agent_info['prenom'])
                    if key not in pool_map:
                        pool_map[key] = agent_info
                        pool.append(agent_info)

            availability[d.strftime('%Y-%m-%d')] = available_agents

        # Assign like generate_week
        assignments = {}
        if len(pool) == 0:
            for d in dates:
                assignments[d.strftime('%Y-%m-%d')] = {'assigned': [], 'available': availability.get(d.strftime('%Y-%m-%d'), [])}
        else:
            cursor = 0
            for d in dates:
                date_key = d.strftime('%Y-%m-%d')
                avail = availability.get(date_key, [])
                picked = []
                if avail:
                    attempts = 0
                    while len(picked) < slots and attempts < len(pool) * 2:
                        candidate = pool[cursor % len(pool)]
                        cursor += 1
                        attempts += 1
                        if any((candidate['nom'] == a['nom'] and candidate['prenom'] == a['prenom']) for a in avail):
                            if not any((candidate['nom'] == p['nom'] and candidate['prenom'] == p['prenom']) for p in picked):
                                picked.append(candidate)
                assignments[date_key] = {'assigned': picked, 'available': availability.get(date_key, [])}

        # Write assigned markers into sheets (mark with 'P')
        written = 0
        for date_key, info in assignments.items():
            d = datetime.strptime(date_key, '%Y-%m-%d')
            sheet_name = f"{['Janvier','Fevrier','Mars','Avril','Mai','Juin','Juillet','Aout','Septembre','Octobre','Novembre','Decembre'][d.month-1]} {d.year}"
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]
            day = d.day
            for agent in info.get('assigned', []):
                row_idx = agent['row']
                col_idx = 15 + day
                # Only write if empty or already 'P'
                current = sheet.cell(row=row_idx, column=col_idx).value
                if current is None or str(current).strip() == '' or str(current).strip().upper() == 'P' or str(current).strip().lower() in ['present','présent']:
                    sheet.cell(row=row_idx, column=col_idx, value='P')
                    written += 1

        wb.save(EXCEL_FILE)

        return jsonify({'success': True, 'written': written, 'backup': backup_file})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reload-excel', methods=['POST'])
def reload_excel():
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({'success': False, 'error': 'Excel file not found'}), 404
        mtime = os.path.getmtime(EXCEL_FILE)
        size = os.path.getsize(EXCEL_FILE)
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheets = wb.sheetnames
        return jsonify({'success': True, 'mtime': mtime, 'size': size, 'sheets': sheets})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/generate-teams', methods=['POST'])
def generate_teams():
    """Génère les équipes et remplit le template Excel"""
    try:
        data = request.json or {}
        week = data.get('week')
        group = data.get('group', 'all')
        team_size = int(data.get('team_size', 3))  # 2 ou 3 agents par équipe

        if not week:
            return jsonify({'success': False, 'error': 'week is required'}), 400

        # Parse semaine ISO
        if '-W' in week:
            parts = week.split('-W')
            year = int(parts[0])
            week_num = int(parts[1])
        else:
            return jsonify({'success': False, 'error': 'Invalid week format'}), 400

        # Calculer les dates (lundi à vendredi seulement - 5 jours)
        import datetime
        jan1 = datetime.date(year, 1, 1)
        week_start = jan1 + datetime.timedelta(weeks=week_num - 1, days=-jan1.weekday())
        dates = [week_start + datetime.timedelta(days=i) for i in range(5)]  # Seulement lundi à vendredi

        # Lire les agents disponibles depuis le fichier principal
        wb_source = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        # Charger les compétences depuis la feuille 'config' (colonnes AE-AH et plus = indices 30+)
        config_sheet = wb_source['config']
        competences_map = {}  # {nom: {chauffeur_pl, macon, aide_macon, enrobé, enginiste, blowpatcher}}
        
        for row_idx in range(3, 76):
            row = [cell.value for cell in config_sheet[row_idx]]
            nom = row[5].strip() if row[5] and isinstance(row[5], str) else row[5]
            if not nom:
                continue
            
            # Colonnes AE-AH (indices 30-33) et plus pour blowpatcher
            competences = {
                'chauffeur_pl': row[30] if len(row) > 30 and row[30] else False,
                'macon': row[31] if len(row) > 31 and row[31] else False,
                'aide_macon': row[32] if len(row) > 32 and row[32] else False,
                'enrobé': row[33] if len(row) > 33 and row[33] else False,
                'enginiste': row[34] if len(row) > 34 and row[34] else False,
                'blowpatcher': row[35] if len(row) > 35 and row[35] else False,
            }
            competences_map[nom] = competences
        
        def in_group(nom, group):
            if group == 'all':
                return True
            nom_upper = nom.upper() if nom else ''
            group_noms = GROUPS.get(group, [])
            return any(gn.upper() in nom_upper or nom_upper in gn.upper() for gn in group_noms)

        absent_statuses = {'CA','RTT','CEX','R','M','AT','F','AST','PC','TP','TAD'}

        # Collecter agents disponibles par jour
        daily_available = {}
        for d in dates:
            month_name = d.strftime('%B %Y')
            month_name_fr = {
                'January': 'Janvier', 'February': 'Fevrier', 'March': 'Mars',
                'April': 'Avril', 'May': 'Mai', 'June': 'Juin',
                'July': 'Juillet', 'August': 'Aout', 'September': 'Septembre',
                'October': 'Octobre', 'November': 'Novembre', 'December': 'Décembre'
            }.get(month_name.split()[0], month_name.split()[0])
            sheet_name = f"{month_name_fr} {d.year}"
            
            available = []
            if sheet_name not in wb_source.sheetnames:
                daily_available[d.strftime('%Y-%m-%d')] = []
                continue

            sheet = wb_source[sheet_name]
            day = d.day
            col_idx = 14 + day

            for row_idx in range(11, 100):
                row = sheet[row_idx]
                nom = row[1].value
                prenom = row[2].value
                
                if not nom or not in_group(nom, group):
                    continue

                cell_value = row[col_idx].value
                status = normalize_status(cell_value)

                if status.upper() not in absent_statuses and (status == '' or status.upper() == 'P'):
                    competences = competences_map.get(nom, {})
                    available.append({
                        'nom': nom,
                        'prenom': prenom,
                        'fullName': f"{nom} {prenom}".strip(),
                        'competences': competences
                    })
            
            daily_available[d.strftime('%Y-%m-%d')] = available

        # Répartir en équipes selon les compétences et le team_size choisi
        def form_smart_teams(agents, team_size):
            """
            Forme des équipes intelligentes basées sur les compétences
            team_size = 2 ou 3 agents par équipe
            """
            if not agents:
                return []
            
            teams = []
            used = set()
            
            # Récupérer les compétences disponibles (tracking par fullName)
            chauffeurs = [a for a in agents if a['competences'].get('chauffeur_pl')]
            macons = [a for a in agents if a['competences'].get('macon')]
            aides = [a for a in agents if a['competences'].get('aide_macon')]
            enrobés = [a for a in agents if a['competences'].get('enrobé')]
            enginistes = [a for a in agents if a['competences'].get('enginiste')]
            blowpatchers = [a for a in agents if a['competences'].get('blowpatcher')]
            
            if team_size == 3:
                # === ÉQUIPES DE 3 ===
                # Priorité 1: Chauffeur PL + Maçon + Aide maçon
                for ch in chauffeurs:
                    if ch['fullName'] in used:
                        continue
                    for mac in macons:
                        if mac['fullName'] in used or mac['fullName'] == ch['fullName']:
                            continue
                        for aide in aides:
                            if aide['fullName'] in used or aide['fullName'] == ch['fullName'] or aide['fullName'] == mac['fullName']:
                                continue
                            teams.append([ch, mac, aide])
                            used.add(ch['fullName'])
                            used.add(mac['fullName'])
                            used.add(aide['fullName'])
                            break
                
                # Priorité 2: Chauffeur PL + Enrobé + Enrobé
                for ch in chauffeurs:
                    if ch['fullName'] in used:
                        continue
                    enr_list = []
                    for enr in enrobés:
                        if enr['fullName'] not in used and enr['fullName'] != ch['fullName']:
                            enr_list.append(enr)
                            if len(enr_list) == 2:
                                break
                    if len(enr_list) == 2:
                        teams.append([ch, enr_list[0], enr_list[1]])
                        used.add(ch['fullName'])
                        used.add(enr_list[0]['fullName'])
                        used.add(enr_list[1]['fullName'])
                
                # Priorité 3: Enginiste + Maçon + Aide maçon
                for eng in enginistes:
                    if eng['fullName'] in used:
                        continue
                    for mac in macons:
                        if mac['fullName'] in used or mac['fullName'] == eng['fullName']:
                            continue
                        for aide in aides:
                            if aide['fullName'] in used or aide['fullName'] == eng['fullName'] or aide['fullName'] == mac['fullName']:
                                continue
                            teams.append([eng, mac, aide])
                            used.add(eng['fullName'])
                            used.add(mac['fullName'])
                            used.add(aide['fullName'])
                            break
                
                # Priorité 4: Maçon + Aide maçon
                for mac in macons:
                    if mac['fullName'] in used:
                        continue
                    for aide in aides:
                        if aide['fullName'] in used or aide['fullName'] == mac['fullName']:
                            continue
                        teams.append([mac, aide])
                        used.add(mac['fullName'])
                        used.add(aide['fullName'])
                        break
            
            elif team_size == 2:
                # === ÉQUIPES DE 2 ===
                # Priorité 1: 2 Blowpatcher
                blop_idx = 0
                while blop_idx < len(blowpatchers) - 1:
                    bp1 = blowpatchers[blop_idx]
                    if bp1['fullName'] in used:
                        blop_idx += 1
                        continue
                    bp2 = blowpatchers[blop_idx + 1]
                    if bp2['fullName'] in used or bp2['fullName'] == bp1['fullName']:
                        blop_idx += 1
                        continue
                    teams.append([bp1, bp2])
                    used.add(bp1['fullName'])
                    used.add(bp2['fullName'])
                    blop_idx += 2
                
                # Priorité 2: Chauffeur PL + Maçon
                for ch in chauffeurs:
                    if ch['fullName'] in used:
                        continue
                    for mac in macons:
                        if mac['fullName'] in used or mac['fullName'] == ch['fullName']:
                            continue
                        teams.append([ch, mac])
                        used.add(ch['fullName'])
                        used.add(mac['fullName'])
                        break
                
                # Priorité 3: Chauffeur PL + Aide maçon
                for ch in chauffeurs:
                    if ch['fullName'] in used:
                        continue
                    for aide in aides:
                        if aide['fullName'] in used or aide['fullName'] == ch['fullName']:
                            continue
                        teams.append([ch, aide])
                        used.add(ch['fullName'])
                        used.add(aide['fullName'])
                        break
            
            # Remplir le reste avec les agents non utilisés
            remaining_agents = [a for a in agents if a['fullName'] not in used]
            
            if remaining_agents:
                if team_size == 3:
                    for i in range(0, len(remaining_agents), 3):
                        teams.append(remaining_agents[i:i+3])
                elif team_size == 2:
                    for i in range(0, len(remaining_agents), 2):
                        teams.append(remaining_agents[i:i+2])
            
            return teams
        
        daily_teams = {}
        for date_key, agents in daily_available.items():
            teams = form_smart_teams(agents, team_size)
            daily_teams[date_key] = teams

        # Créer un nouveau workbook pour le planning
        output_file = f"planning_semaine_{week}.xlsx"
        wb_template = openpyxl.Workbook()
        
        # S'assurer qu'on a une feuille active
        sheet = wb_template.active
        if sheet is None:
            sheet = wb_template.create_sheet()
        
        # Vérifier que sheet n'est pas None
        if sheet is None:
            return jsonify({'error': 'Impossible de créer la feuille'}), 500
        
        # Remplir le tableau : 5 colonnes (DATE | AGENT 1 | AGENT 2 | AGENT 3 | INTERVENTION)
        # En-têtes en ligne 1
        sheet.cell(row=1, column=1, value='DATE')
        sheet.cell(row=1, column=2, value='EQUIPE')
        sheet.cell(row=1, column=3, value='')
        sheet.cell(row=1, column=4, value='')
        sheet.cell(row=1, column=5, value='INTERVENTION')
        
        # Styles pour en-têtes
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col in [1, 2, 5]:
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Police 10 par défaut pour tout le document
        default_font = Font(size=10)
        
        # Couleurs alternées pour les jours (sobres)
        day_colors = [
            'E8F4F8',  # Bleu très clair
            'F0F0F0',  # Gris très clair
        ]
        
        day_names_fr = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']
        
        current_row = 2
        for day_idx, (d, day_name) in enumerate(zip(dates, day_names_fr)):
            date_key = d.strftime('%Y-%m-%d')
            teams = daily_teams.get(date_key, [])
            
            # Couleur alternée pour ce jour
            day_fill = PatternFill(start_color=day_colors[day_idx % 2], end_color=day_colors[day_idx % 2], fill_type='solid')
            
            if len(teams) == 0:
                # Aucune équipe : une ligne vide
                sheet.cell(row=current_row, column=1, value=f"{day_name} {d.strftime('%d/%m/%Y')}")
                sheet.cell(row=current_row, column=2, value='')
                sheet.cell(row=current_row, column=3, value='')
                sheet.cell(row=current_row, column=4, value='')
                sheet.cell(row=current_row, column=5, value='')
                # Colonne 1 (DATE) alignée à gauche et en haut
                cell = sheet.cell(row=current_row, column=1)
                cell.fill = day_fill
                cell.font = default_font
                cell.alignment = Alignment(horizontal='left', vertical='top')
                # Colonnes 2-5 centrées
                for col in [2, 3, 4, 5]:
                    cell = sheet.cell(row=current_row, column=col)
                    cell.fill = day_fill
                    cell.font = default_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
            else:
                # Une ligne par équipe
                for team_idx, team in enumerate(teams):
                    # Première ligne du jour : afficher la date
                    if team_idx == 0:
                        sheet.cell(row=current_row, column=1, value=f"{day_name} {d.strftime('%d/%m/%Y')}")
                    else:
                        sheet.cell(row=current_row, column=1, value='')
                    
                    # Remplir les agents de cette équipe (respecter team_size)
                    for agent_idx, agent in enumerate(team):
                        if agent_idx < 3:  # Max 3 colonnes d'agents (colonnes 2, 3, 4)
                            sheet.cell(row=current_row, column=2+agent_idx, value=agent['fullName'])
                        else:
                            break  # Ignorer si > 3 agents
                    
                    # Laisser vides les colonnes non utilisées (si team_size < 3)
                    for unused_col in range(len(team), 3):
                        sheet.cell(row=current_row, column=2+unused_col, value='')
                    
                    sheet.cell(row=current_row, column=5, value='')
                    
                    # Colonne 1 (DATE) alignée à gauche et en haut
                    cell = sheet.cell(row=current_row, column=1)
                    cell.fill = day_fill
                    cell.font = default_font
                    cell.alignment = Alignment(horizontal='left', vertical='top')
                    # Colonnes 2-5 centrées
                    for col in [2, 3, 4, 5]:
                        cell = sheet.cell(row=current_row, column=col)
                        cell.fill = day_fill
                        cell.font = default_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    current_row += 1
        
        # Ajuster les largeurs de colonnes selon le template
        sheet.column_dimensions['A'].width = 17.0
        sheet.column_dimensions['B'].width = 20.140625
        sheet.column_dimensions['C'].width = 20.140625
        sheet.column_dimensions['D'].width = 20.140625
        sheet.column_dimensions['E'].width = 20.140625
        
        # Configuration de l'impression : marges étroites et quadrillage
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.print_options.gridLines = True
        sheet.print_options.gridLinesSet = True
        
        # Marges étroites (en inches)
        sheet.page_margins.left = 0.25
        sheet.page_margins.right = 0.25
        sheet.page_margins.top = 0.75
        sheet.page_margins.bottom = 0.75
        sheet.page_margins.header = 0.3
        sheet.page_margins.footer = 0.3
        
        # Sauvegarder en mémoire et renvoyer comme fichier téléchargeable
        output = BytesIO()
        wb_template.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_file
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/load-planning', methods=['POST'])
def load_planning():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Aucun fichier fourni'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Nom de fichier vide'}), 400
        
        # Charger le fichier Excel
        wb = openpyxl.load_workbook(file.stream, data_only=True)
        
        # Prendre la première feuille
        sheet = wb.active
        if sheet is None:
            if len(wb.worksheets) == 0:
                return jsonify({'success': False, 'error': 'Le fichier ne contient aucune feuille'}), 400
            sheet = wb.worksheets[0]
        
        # Lire les en-têtes (première ligne non vide)
        headers = []
        first_row = None
        for row_idx in range(1, min(10, sheet.max_row + 1)):
            row_values = [cell.value for cell in sheet[row_idx]]
            if any(v for v in row_values):
                headers = [str(v) if v else '' for v in row_values]
                first_row = row_idx
                break
        
        if not headers or first_row is None:
            return jsonify({'success': False, 'error': 'Aucun en-tête trouvé'}), 400
        
        # Lire toutes les lignes de données
        rows = []
        for row_idx in range(first_row + 1, sheet.max_row + 1):
            row_values = [cell.value for cell in sheet[row_idx]]
            # Ignorer les lignes complètement vides
            if any(v for v in row_values):
                rows.append([str(v) if v else '' for v in row_values])
        
        return jsonify({
            'success': True,
            'title': f'Planning - {file.filename}',
            'headers': headers,
            'rows': rows
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download-excel', methods=['GET'])
def download_excel():
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({'success': False, 'error': 'Fichier introuvable'}), 404
        
        return send_file(
            EXCEL_FILE,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='2026 - PRESENCES_CONGES VOIRIE ESPACES VERTS ST8.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
