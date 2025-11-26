import openpyxl
wb=openpyxl.load_workbook('2026 - PRESENCES_CONGES VOIRIE ESPACES VERTS ST8 (1).xlsx', data_only=True)
s=wb['config']
headers=[cell.value for cell in s[2]]
print('Headers cols 1..40:')
for i,h in enumerate(headers[:40], start=1):
    print(i, h)
print('\nAgents with non-empty aptitude cols (cols 8-30):')
for r in range(3,76):
    nom=s.cell(r,6).value
    pren=s.cell(r,7).value
    if not nom:
        continue
    entries=[]
    for c in range(8,31):
        val=s.cell(r,c).value
        if val not in (None,''):
            header=headers[c-1] if c-1<len(headers) else f'col{c}'
            entries.append((c, header, val))
    if entries:
        print(f"\nRow {r}: {nom} {pren}")
        for e in entries:
            print(' ', e)
