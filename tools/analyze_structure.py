import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Path to the file
XLSX = r'C:\Users\Gonçalves\Desktop\st88\2026 - PRESENCES_CONGES VOIRIE ESPACES VERTS ST8 (1).xlsx'

print('='*80)
print('ANALYSE DU FICHIER:', XLSX)
print('='*80)

if not os.path.exists(XLSX):
    print('ERREUR: Fichier introuvable:', XLSX)
    exit(1)

print(f'\nTaille: {os.path.getsize(XLSX)} octets')

# Load workbook (read-only to avoid any modification)
wb = load_workbook(filename=XLSX, data_only=False, read_only=False)

print(f'\nNombre de feuilles: {len(wb.sheetnames)}')
print('Feuilles:', ', '.join(wb.sheetnames))

# Analyze each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print('\n' + '='*80)
    print(f'FEUILLE: {sheet_name}')
    print('='*80)
    
    # Dimensions
    print(f'Dimensions: {ws.dimensions}')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    
    # Merged cells
    merged = list(ws.merged_cells.ranges)
    if merged:
        print(f'\nCellules fusionnées ({len(merged)}):')
        for mr in merged[:10]:  # show first 10
            print(f'  - {mr}')
        if len(merged) > 10:
            print(f'  ... et {len(merged)-10} autres')
    else:
        print('\nAucune cellule fusionnée')
    
    # First 8 rows (headers + sample data)
    print('\n8 premières lignes:')
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        # Show first 12 columns
        row_preview = row[:12] if len(row) > 12 else row
        row_str = ' | '.join([str(c)[:20] if c is not None else '' for c in row_preview])
        print(f'  Row {r_idx}: {row_str}')
        if len(row) > 12:
            print(f'           ... + {len(row)-12} colonnes supplémentaires')
    
    # Column widths
    print('\nLargeurs de colonnes (10 premières):')
    for col_idx in range(1, min(11, ws.max_column + 1)):
        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width
        print(f'  Col {col_letter}: {width}')
    
    # Row heights
    print('\nHauteurs de lignes (10 premières):')
    for row_idx in range(1, min(11, ws.max_row + 1)):
        height = ws.row_dimensions[row_idx].height
        print(f'  Row {row_idx}: {height}')
    
    # Check for formulas
    formula_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row)):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
    if formula_count > 0:
        print(f'\nFormules détectées: {formula_count} (dans les 50 premières lignes)')
    
    # Sample cell styles (first data cell)
    if ws.max_row > 1:
        sample_cell = ws.cell(row=2, column=1)
        print(f'\nStyle de la cellule A2:')
        print(f'  Font: {sample_cell.font.name}, taille={sample_cell.font.size}, gras={sample_cell.font.bold}')
        print(f'  Fill: {sample_cell.fill.patternType}')
        print(f'  Alignment: horizontal={sample_cell.alignment.horizontal}, vertical={sample_cell.alignment.vertical}')
        print(f'  Border: {sample_cell.border.left.style if sample_cell.border.left else None}')

print('\n' + '='*80)
print('ANALYSE TERMINÉE')
print('='*80)
