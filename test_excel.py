import openpyxl

wb = openpyxl.load_workbook('2026 - PRESENCES_CONGES VOIRIE ESPACES VERTS ST8 (1).xlsx', data_only=True)
sheet = wb['Janvier 2026']

agents = []
for r in range(11, 100):
    nom = sheet.cell(row=r, column=2).value
    prenom = sheet.cell(row=r, column=3).value
    if nom:
        agents.append((nom, prenom))

print(f"Total agents dans Excel: {len(agents)}\n")
for i, (nom, prenom) in enumerate(agents, 1):
    print(f"{i}. {nom!r} {prenom!r}")
